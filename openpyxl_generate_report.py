import os
import json
import boto3
import openpyxl
import openpyxl
from botocore.exceptions import ClientError
import datetime
import logging
import pytz
import tzlocal
import dateutil.parser
from pprint import pprint

class Report():
  wb = openpyxl.load_workbook('inventory.xlsx')
  wb.sheetnames
  sheet = wb.active
  currSheet = sheet.title
  def __init__(self,acct,creation_date,rsrc_type,rsrc_name,owner):
    self.acct = acct
    self.creation_date = creation_date
    self.rsrc_type = rsrc_type
    self.rsrc_name = rsrc_name
    self.owner = owner

  def write_report(self):
    wb = openpyxl.load_workbook('inventory.xlsx')
    wb.sheetnames
    sheet = wb.active
    currSheet = sheet.title
    sheet.append((self.acct,self.creation_date,self.rsrc_type,self.rsrc_name,self.owner))
    wb.save('inventory.xlsx') 